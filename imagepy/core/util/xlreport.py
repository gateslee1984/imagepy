import openpyxl as pyxl
from openpyxl.utils.units import cm_to_EMU, EMU_to_pixels
from io import BytesIO
from openpyxl.drawing.image import Image
from PIL import Image as PImage
import numpy as np
import pandas as pd
from copy import copy

if not '.rpt' in pyxl.reader.excel.SUPPORTED_FORMATS:
    pyxl.reader.excel.SUPPORTED_FORMATS += ('.rpt',)

def parse(wb):
    rst = []
    for ws in wb.worksheets:
        rst.append((ws.title, []))
        for row in ws.rows:
            for cell in row:
                if not isinstance(cell.value, str):continue
                if cell.value[0]+cell.value[-1] != '{}': continue
                rst[-1][-1].append(((cell.row, cell.col_idx),
                            cell.value[1:-1].split(' ')))
    return rst

def trans(img, W, H, margin, scale):
    h, w = img.shape[:2]
    h2, w2 = int(h/margin), int(w/margin)
    if scale:
        if W/H > w/h: w2 = int(W/H*h2)
        if H/W > h/w: h2 = int(H/W*w2)
    newshp = (h2, w2) if img.ndim==2 else (h2, w2, 3)
    blank = np.ones(newshp, dtype=np.uint8) * 255
    blank[(h2-h)//2:(h2-h)//2+h, (w2-w)//2:(w2-w)//2+w] = img
    return blank

def add_image(wb, ws, pos, key, img):
    if img is None: return
    w, h = [float(i) for i in key[2].split(':')]
    margin, scale = [float(i) for i in key[3].split(':')]
    img = trans(img, w, h, margin, scale==0)
    
    img = PImage.fromarray(img)
    image_file = BytesIO()
    img.save(image_file, 'png')
    ref = BytesIO(image_file.getvalue())
    image = Image(img)
    image.ref = ref
    image.height = EMU_to_pixels(cm_to_EMU(h))
    image.width = EMU_to_pixels(cm_to_EMU(w))
    wb[ws].add_image(image, wb[ws].cell(*pos).coordinate)
                
def add_table(wb, ws, pos, key, data):
    if data is None: return
    vs = data.values
    idx, cols = data.index, data.columns
    dr, dc = [int(i) for i in key[2].split(':')] if len(key)>2 else (1,1)
    ir, ic = [int(i) for i in key[3].split(':')] if len(key)>3 else (0,0)
    for r in range(vs.shape[0]):
        if ir!=0: wb[ws].cell(pos[0]+r*dr, pos[1]+ir, idx[r])
    for c in range(vs.shape[1]):
        if ic!=0: wb[ws].cell(pos[0]+ic, pos[1]+c*dc, cols[c])
    for r in range(vs.shape[0]):
        for c in range(vs.shape[1]):
            wb[ws].cell(pos[0]+r*dr, pos[1]+c*dc, vs[r,c])

def fill_value(wb, infos, para):
    for worksheet in infos:
        ws, info = worksheet
        for pos, key in info:
            if not key[1] in para: continue
            if key[0] in ('str', 'int', 'float', 'bool', 'txt', 'list', 'date'):
                wb[ws].cell(pos[0], pos[1], para[key[1]])
            if key[0] == 'img':
                add_image(wb, ws, pos, key, para[key[1]])
            if key[0] == 'tab':
                add_table(wb, ws, pos, key, para[key[1]])
                

def repair(wb):
    for ws in wb.worksheets:
        for cr in ws.merged_cells:
            ltc = ws.cell(cr.min_row, cr.min_col)
            vb, hb = ltc.border.left, ltc.border.top
            for r in range(cr.min_row, cr.max_row+1):
                for c in range(cr.min_col, cr.max_col+1):
                    cur = copy(ws.cell(r, c).border)
                    cur.left, cur.right = copy(vb), copy(vb)
                    cur.top, cur.bottom = copy(hb), copy(hb)
                    ws.cell(r, c).border = cur

if __name__ == '__main__':
    rst = pd.read_csv('rst.csv')
    img = np.arange(10000, dtype=np.uint8).reshape((100,100))
    data = {'id':'Coins-0001', 'name':'YX Dragon', 'date':'2019-02-05',
            'data':rst, 'ori':img, 'msk':img}
        
    wb = pyxl.load_workbook('temp.xlsx',)
    repair(wb)
    ws = wb.active
    
    
    infos = parse(wb)
    fill_value(wb, infos, data)
    wb.save('new.xlsx')
    
